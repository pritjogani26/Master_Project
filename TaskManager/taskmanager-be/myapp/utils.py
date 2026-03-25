from django.db import connection

import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment



def get_user_role(user_id):

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT r.name
            FROM roles r
            JOIN user_roles ur ON r.id = ur.role_id
            WHERE ur.user_id = %s
        """, [user_id])

        row = cursor.fetchone()

    return row[0] if row else None



def create_login_activity_pdf(data_list):
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Login Activity Report</b>", styles['Title']))
    elements.append(Spacer(1, 20))

    headers = ["ID", "Email", "IP Address", "Status", "Login Time", "Active"]
    table_data = [headers]

    for row in data_list:
        raw_time = row.get("login_time", "")
        # Safely cast to string so it works with both DB Datetime objects and JSON strings
        clean_time = str(raw_time)[:19].replace('T', ' ') if raw_time else "N/A"
        
        table_data.append([
            str(row.get("id", "")),
            str(row.get("email", "")),
            str(row.get("ip_address", "")),
            str(row.get("status", "")),
            clean_time,
            "Yes" if row.get("is_active") else "No"
        ])

    # Create the Table object
    col_widths = [40, 150, 100, 80, 150, 60]
    t = Table(table_data, colWidths=col_widths)
    
    # Apply professional styling
    t.setStyle(TableStyle([
        # Header Row Style
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Data Rows Style
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F8F9F9")),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#BDC3C7")),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        
        ('ALIGN', (1, 0), (1, -1), 'LEFT'), 
    ]))

    # Build the PDF into the buffer
    elements.append(t)
    doc.build(elements)

    # 3. Rewind the buffer to the beginning so Django can read it
    buffer.seek(0)
    
    # 4. Return the buffer to the View
    return buffer




def create_login_activity_excel(data_list):
    # 1. Create a new Workbook in memory and get the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Login Activity"

    # 2. Define and append the headers
    headers = ["ID", "Email", "IP Address", "Status", "Login Time", "Active"]
    ws.append(headers)

    # Optional: Make the header row look professional (Dark blue background, white bold text)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Widen the columns slightly so the data isn't squished
        ws.column_dimensions[cell.column_letter].width = 20

    # 3. Loop through your data and append the rows
    for row in data_list:
        raw_time = row.get("login_time", "")
        # The same safe datetime-to-string fix we used for the PDF
        clean_time = str(raw_time)[:19].replace('T', ' ') if raw_time else "N/A"
        
        ws.append([
            row.get("id", ""),
            row.get("email", ""),
            row.get("ip_address", ""),
            row.get("status", ""),
            clean_time,
            "Yes" if row.get("is_active") else "No"
        ])

    # 4. Save the workbook to our RAM buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    
    # Rewind the buffer
    buffer.seek(0)
    
    return buffer


